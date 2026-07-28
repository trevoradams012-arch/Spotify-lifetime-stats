#!/usr/bin/env python3
"""Parse Spotify Extended Streaming History JSON into a sortable XLSX.

Reads every Streaming_History_Audio_*.json in a directory, drops plays shorter
than a threshold (20 seconds by default), and writes a workbook with one sheet
each for songs, artists, per-year totals, and an overall summary.

Field names follow ReadMeFirst_ExtendedStreamingHistory.pdf:
  ts                                 stream end time, UTC
  ms_played                          milliseconds the stream was played
  master_metadata_track_name         track name (null for podcasts/audiobooks)
  master_metadata_album_artist_name  artist name
  master_metadata_album_album_name   album name

Usage:
  python3 spotify_stats.py [--min-seconds 20] [--indir .] [--out FILE.xlsx]
"""

import argparse
import glob
import json
import os
import re
from collections import Counter, defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1DB954")  # Spotify green
MS_PER_HOUR = 3_600_000
MS_PER_MIN = 60_000

# --- Title variant merging ------------------------------------------------
#
# A Spotify title carries edition info in trailing segments, either after " - "
# or inside parentheses: "1979 - Remastered 2012", "Song (feat. Drake)".
#
# The rule below is: merge segments that mark the SAME STUDIO RECORDING in a
# different edition, master, or trim. Keep anything that marks a DIFFERENT
# PERFORMANCE (live, acoustic, demo, re-recording) as its own song.
#
# Anything not recognized is kept, so an unknown suffix never silently merges.

# Any title containing "remix" is left completely untouched, per the rule that
# nothing with remix in it gets merged. Checked before all other patterns.
REMIX_RE = re.compile(r"re-?mix", re.I)

# Segments dropped from the key: same recording, different edition/master/trim.
MERGE_SEGMENT_RES = [re.compile(p, re.I) for p in (
    r"(19|20)\d{2}\s*-?\s*(digital\s+)?remaster(ed)?(\s+version)?",
    r"(digitally\s+|digital\s+)?remaster(ed)?(\s+(19|20)\d{2})?(\s+version)?",
    r"(the\s+)?album\s+version.*",
    r"original(\s+(mix|version|single\s+version))?",
    r"single\s+(version|edit|mix)",
    r"radio\s+(edit|version|mix)",
    r"edit(ed)?",
    r"bonus\s+track.*",
    r"ep\s+version",
    r"(non-?)?lp\s+version",
    r"main\s+version",
    r"full[- ]length\s+version",
    r"(mono|stereo)(\s+(version|mix))?",
    r"deluxe(\s+(edition|version))?",
    r"re-?issue",
    r"(explicit|clean)(\s+version)?",
    r"feat\..*", r"featuring\s+.*", r"ft\..*",
)]

# Segments that force a separate song even though a pattern above might match
# part of them. Checked before MERGE_SEGMENT_RES.
KEEP_SEGMENT_RES = [re.compile(p, re.I) for p in (
    r".*\blive\b.*", r".*\bsession[s]?\b.*", r".*\bunplugged\b.*",
    r".*\bacoustic\b.*", r".*\binstrumental\b.*", r".*\bdemo\b.*",
    r".*\bcommentary\b.*", r".*\bcover\b.*", r".*\bkaraoke\b.*",
    r".*\bre-?record(ed|ing)?\b.*", r".*\btaylor.s version\b.*",
    r".*\breimagined\b.*", r".*\brevisited\b.*", r".*\bremake\b.*",
    r".*\balternat(e|ive)\b.*", r".*\bextended\b.*", r".*\bclub\b.*",
    r".*\bdub\b.*", r".*\bvip\b.*", r".*\bbootleg\b.*", r".*\brework\b.*",
)]

PAREN_RE = re.compile(r"\s*[\(\[]([^)\]]*)[\)\]]\s*")


def merge_segment(seg):
    """True if this trailing segment marks an edition rather than a new song."""
    seg = seg.strip()
    if not seg:
        return False
    if any(r.fullmatch(seg) for r in KEEP_SEGMENT_RES):
        return False
    return any(r.fullmatch(seg) for r in MERGE_SEGMENT_RES)


def canonical_title(title):
    """Strip edition-only suffixes so variants of one recording share a key.

    "1979 - Remastered 2012"      -> "1979"
    "Song (feat. Drake)"          -> "Song"
    "Song - Live - 2011 Remaster" -> "Song - Live"   (live stays distinct)
    "Song - Skrillex Remix"       -> unchanged       (never merged)
    """
    if REMIX_RE.search(title):
        return title.strip()

    # Parentheticals can appear anywhere; keep the ones that aren't edition marks.
    kept_parens = []

    def take(match):
        inner = match.group(1).strip()
        if not merge_segment(inner):
            kept_parens.append(inner)
        return " "

    stripped = PAREN_RE.sub(take, title).strip()

    parts = [p.strip() for p in stripped.split(" - ")]
    base, tail = parts[0], parts[1:]
    tail = [seg for seg in tail if not merge_segment(seg)]

    key = " - ".join([base] + tail)
    for inner in kept_parens:
        key += f" ({inner})"
    return key.strip() or title.strip()


def load_streams(indir):
    """Yield every audio stream record from the export."""
    paths = sorted(glob.glob(os.path.join(indir, "Streaming_History_Audio_*.json")))
    if not paths:
        raise SystemExit(f"No Streaming_History_Audio_*.json files found in {indir}")
    for path in paths:
        with open(path, encoding="utf-8") as fh:
            for rec in json.load(fh):
                yield rec
    print(f"Read {len(paths)} JSON files")


def aggregate(indir, min_ms):
    """Roll streams up by song and by artist, skipping short plays."""
    songs = defaultdict(lambda: {
        "plays": 0, "ms": 0, "names": Counter(), "artists": Counter(),
        "albums": Counter(), "first": None, "last": None, "uri": None,
    })
    artists = defaultdict(lambda: {
        "plays": 0, "ms": 0, "names": Counter(), "tracks": set(),
        "first": None, "last": None,
    })
    years = defaultdict(lambda: {"plays": 0, "ms": 0, "tracks": set(), "artists": set()})

    total = kept = skipped_short = non_music = 0

    for rec in load_streams(indir):
        total += 1
        track = rec.get("master_metadata_track_name")
        artist = rec.get("master_metadata_album_artist_name")
        if not track or not artist:
            non_music += 1  # podcast episode or audiobook chapter
            continue
        ms = rec.get("ms_played") or 0
        if ms < min_ms:
            skipped_short += 1
            continue
        kept += 1

        ts = rec.get("ts") or ""
        # ISO-8601 timestamps sort lexically, so min/max need no date parsing.
        date = ts[:10]
        year = ts[:4]
        album = rec.get("master_metadata_album_album_name") or ""

        # Case-insensitive keys, with edition-only suffixes stripped, so that
        # "Song", "song", and "Song - 2012 Remaster" collapse into one row.
        song_key = (canonical_title(track).casefold(), artist.strip().casefold())
        artist_key = artist.strip().casefold()

        s = songs[song_key]
        s["plays"] += 1
        s["ms"] += ms
        s["names"][track] += 1
        s["artists"][artist] += 1
        if album:
            s["albums"][album] += 1
        s["uri"] = s["uri"] or rec.get("spotify_track_uri")
        if date:
            s["first"] = min(s["first"] or date, date)
            s["last"] = max(s["last"] or date, date)

        a = artists[artist_key]
        a["plays"] += 1
        a["ms"] += ms
        a["names"][artist] += 1
        a["tracks"].add(song_key)
        if date:
            a["first"] = min(a["first"] or date, date)
            a["last"] = max(a["last"] or date, date)

        if year:
            y = years[year]
            y["plays"] += 1
            y["ms"] += ms
            y["tracks"].add(song_key)
            y["artists"].add(artist_key)

    stats = {
        "total": total, "kept": kept,
        "skipped_short": skipped_short, "non_music": non_music,
    }
    return songs, artists, years, stats


def style_header(ws, ncols):
    """Bold green header row, frozen, with an autofilter for sorting."""
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ws.max_row}"


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_songs(wb, songs):
    ws = wb.create_sheet("Top Songs")
    headers = ["Rank", "Song", "Artist", "Album", "Hours Played",
               "Minutes Played", "Plays", "Avg Minutes/Play", "Titles Merged",
               "First Played", "Last Played", "Total ms Played"]
    ws.append(headers)

    # Ranked by listening time; play count breaks ties.
    rows = sorted(songs.values(), key=lambda s: (-s["ms"], -s["plays"]))
    for i, s in enumerate(rows, start=1):
        ws.append([
            i,
            s["names"].most_common(1)[0][0],
            s["artists"].most_common(1)[0][0],
            s["albums"].most_common(1)[0][0] if s["albums"] else "",
            s["ms"] / MS_PER_HOUR,
            s["ms"] / MS_PER_MIN,
            s["plays"],
            s["ms"] / s["plays"] / MS_PER_MIN,
            len(s["names"]),
            s["first"] or "",
            s["last"] or "",
            s["ms"],
        ])
    finish(ws, headers=headers,
           widths=[6, 42, 28, 34, 12, 13, 8, 13, 11, 12, 12, 14],
           num_formats={5: "#,##0.00", 6: "#,##0.0", 8: "#,##0.00",
                        7: "#,##0", 9: "#,##0", 12: "#,##0"})
    return len(rows)


def write_merge_audit(wb, songs):
    """One row per raw Spotify title that was folded into another, for review."""
    ws = wb.create_sheet("Merge Audit")
    headers = ["Song (kept name)", "Artist", "Hours Played", "Total Plays",
               "Titles Merged", "Merged Spotify Titles (plays each)"]
    ws.append(headers)

    merged = [s for s in songs.values() if len(s["names"]) > 1]
    merged.sort(key=lambda s: (-s["ms"], -s["plays"]))
    for s in merged:
        variants = "  |  ".join(f"{n} ({c})" for n, c in s["names"].most_common())
        ws.append([
            s["names"].most_common(1)[0][0],
            s["artists"].most_common(1)[0][0],
            s["ms"] / MS_PER_HOUR,
            s["plays"],
            len(s["names"]),
            variants,
        ])
    finish(ws, headers=headers, widths=[42, 28, 12, 11, 12, 110],
           num_formats={3: "#,##0.00", 4: "#,##0", 5: "#,##0"})
    return len(merged)


def write_artists(wb, artists):
    ws = wb.create_sheet("Top Artists")
    headers = ["Rank", "Artist", "Hours Played", "Minutes Played", "Plays",
               "Distinct Songs", "Avg Minutes/Play", "First Played", "Last Played",
               "Total ms Played"]
    ws.append(headers)

    # Ranked by listening time; play count breaks ties.
    rows = sorted(artists.values(), key=lambda a: (-a["ms"], -a["plays"]))
    for i, a in enumerate(rows, start=1):
        ws.append([
            i,
            a["names"].most_common(1)[0][0],
            a["ms"] / MS_PER_HOUR,
            a["ms"] / MS_PER_MIN,
            a["plays"],
            len(a["tracks"]),
            a["ms"] / a["plays"] / MS_PER_MIN,
            a["first"] or "",
            a["last"] or "",
            a["ms"],
        ])
    finish(ws, headers=headers,
           widths=[6, 34, 12, 13, 8, 13, 13, 12, 12, 14],
           num_formats={3: "#,##0.00", 4: "#,##0.0", 7: "#,##0.00",
                        5: "#,##0", 6: "#,##0", 10: "#,##0"})
    return len(rows)


def write_years(wb, years):
    ws = wb.create_sheet("By Year")
    headers = ["Year", "Plays", "Hours Played", "Distinct Songs", "Distinct Artists",
               "Total ms Played"]
    ws.append(headers)
    for year in sorted(years):
        y = years[year]
        ws.append([year, y["plays"], y["ms"] / MS_PER_HOUR,
                   len(y["tracks"]), len(y["artists"]), y["ms"]])
    total_plays = sum(y["plays"] for y in years.values())
    total_ms = sum(y["ms"] for y in years.values())
    ws.append(["Total", total_plays, total_ms / MS_PER_HOUR, "", "", total_ms])
    total_row = ws.max_row
    finish(ws, headers=headers,
           widths=[10, 12, 13, 14, 15, 16],
           num_formats={2: "#,##0", 3: "#,##0.0", 4: "#,##0", 5: "#,##0", 6: "#,##0"},
           autofilter=False)
    for col in range(1, len(headers) + 1):
        ws.cell(row=total_row, column=col).font = Font(name=FONT, bold=True)
    return total_ms


def write_summary(wb, stats, n_songs, n_artists, n_merged, total_ms, min_seconds):
    ws = wb.create_sheet("Summary", 0)
    ws.append(["Spotify Extended Streaming History", ""])
    ws["A1"].font = Font(name=FONT, bold=True, size=14)

    rows = [
        ("", ""),
        ("Minimum play length counted (seconds)", min_seconds),
        ("", ""),
        ("Streams in export (all audio files)", stats["total"]),
        ("  Podcast / audiobook streams excluded", stats["non_music"]),
        (f"  Music plays under {min_seconds:g}s excluded", stats["skipped_short"]),
        ("Music plays counted", stats["kept"]),
        ("", ""),
        ("Distinct songs", n_songs),
        ("  Songs with merged title variants", n_merged),
        ("Distinct artists", n_artists),
        ("Total hours listened", total_ms / MS_PER_HOUR),
        ("Total days listened", total_ms / MS_PER_HOUR / 24),
        ("", ""),
        ("Remasters, album/single/radio versions and (feat. X) titles are", ""),
        ("merged into one song. Live, acoustic, instrumental, demo and", ""),
        ("anything with 'remix' stay separate. See the Merge Audit sheet.", ""),
        ("", ""),
        ("Ranked by total time played. Every sheet has filter arrows in", ""),
        ("its header row - click one and choose Sort Largest to Smallest", ""),
        ("to re-rank by Plays or any other column.", ""),
    ]
    for label, value in rows:
        ws.append([label, value])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=2):
        for cell in row:
            cell.font = Font(name=FONT)
            if isinstance(cell.value, int):
                cell.number_format = "#,##0"
            elif isinstance(cell.value, float):
                cell.number_format = "#,##0.0"
    for r in range(ws.max_row - 6, ws.max_row + 1):
        ws.cell(row=r, column=1).font = Font(name=FONT, italic=True, size=9)
    set_widths(ws, [58, 18])


def finish(ws, headers, widths, num_formats, autofilter=True):
    """Apply font, number formats, widths, and header styling to a data sheet."""
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.font = Font(name=FONT)
            fmt = num_formats.get(cell.column)
            if fmt:
                cell.number_format = fmt
    set_widths(ws, widths)
    style_header(ws, len(headers))
    if not autofilter:
        ws.auto_filter.ref = None


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--indir", default=".", help="directory holding the JSON files")
    ap.add_argument("--out", default="Spotify_Listening_Stats.xlsx", help="output .xlsx")
    ap.add_argument("--min-seconds", type=float, default=20.0,
                    help="drop plays shorter than this (default: 20)")
    args = ap.parse_args()

    min_ms = int(args.min_seconds * 1000)
    songs, artists, years, stats = aggregate(args.indir, min_ms)

    print(f"{stats['total']:,} streams -> {stats['kept']:,} counted "
          f"({stats['skipped_short']:,} under {args.min_seconds:g}s, "
          f"{stats['non_music']:,} non-music)")
    print(f"{len(songs):,} distinct songs, {len(artists):,} distinct artists")

    wb = Workbook()
    wb.remove(wb.active)
    n_songs = write_songs(wb, songs)
    n_artists = write_artists(wb, artists)
    total_ms = write_years(wb, years)
    n_merged = write_merge_audit(wb, songs)
    write_summary(wb, stats, n_songs, n_artists, n_merged, total_ms, args.min_seconds)
    print(f"{n_merged:,} songs had title variants merged")

    out = args.out if os.path.isabs(args.out) else os.path.join(args.indir, args.out)
    wb.save(out)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
